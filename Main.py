import openpyxl
#Theory: casual relationships dont care about political standpoint
#For this, i'm not going to include it in the algorithm and see how it goes
# (its a trial test run)


total_rows = int(input("What is the last row of the excel spreadsheet?: "))
total_col = 43

#Importing Excel spreadsheet

# Load the Excel workbook
workbook = openpyxl.load_workbook('Campus_SV_New.xlsx')
# Select the worksheet you want to work with
worksheet = workbook['Form Responses 1']

#Main List of all array clients
main_List = []
object = []

#For Loop for creating an array for every client and assigning its values from excel

#cell_value = worksheet.cell(row=row_number, column=column_number).value

for row in range(2, total_rows + 1):
    for colm in range(3, total_col + 1):
        client_ans = worksheet.cell(row, colm).value
        object.append(client_ans)
    #Adding space for match rate
    object.append(0.0)
    main_List.append(object)
    #Resets
    object = []

#Test
#print(main_List[1])
#print(len(main_List[1]))
#x = int(input("Stopper"))

#For Loop for assigning all clients to their list
#Note: Need to organize based on how important you want your rel status to be

#All Lists

gay_list = [] #Add bisexual males

lesb_list = [] #Add bisexual females

straight_cas_list_m = []
straight_cas_list_f = [] #Adding bisexual females

#Side-Note: there's an abundance of straigh_cas_males, so no need to add bi males in the list
straight_ser_list_m = []
straight_ser_list_f = []

#index 3 is sex
#index 4 is sex attracted to

#index 5 is type of rel
#index 6 is the imp of that rel

for client in main_List:
    #LGBTQ+:
    if(client[3] == 'M') and (client[4] == 'M'): #Gays
        gay_list.append(client)
    elif(client[3] == 'F') and (client[4] == 'F'): #Lesbians
        lesb_list.append(client)

    elif(client[3] == 'F') and (client[4] == 'A'): #Bi Girls
        if(client[5] == 'DCA'): #if casual
            straight_cas_list_f.append(client)
        else:
            lesb_list.append(client) #if not, to lesb
    elif(client[3] == 'M') and (client[4] == 'A'): #Bi guys
        gay_list.append(client)

    #Straight:
    elif(client[3] == 'F') and (client[4] == 'M'):
        if (client[5] == 'DCA'):
            straight_cas_list_f.append(client)
        else:
            straight_ser_list_f.append(client)
    elif(client[3] == 'M') and (client[4] == 'F'):
        if (client[5] == 'DCA'):
            straight_cas_list_m.append(client)
        else:
            straight_ser_list_m.append(client)



Preg_Fund = int(input("Give me the percentage the Preguntas_Fundamentales are going to take in the algorithm: "))
Preg_Esp = 100 - Preg_Fund


def print_Res(client):
    print("")
    print("Client ", client[0], ", ", client[3],"4",client[4], ". WhatsApp: ", client[2])
    print("Results: ")

def print_Partner(client, partner):
    #This solves for the same-sex lists
    if(client != partner):
        print("Partner: ", partner[0], " Social: ", partner[1], " Match Rate: ", round(partner[41], 2), "% ")


# I need to find a way to identify the index of every girl when rearranging the dam All_scores
        # Solution: Every client will have the match rate at the end of their array, that way they are
        # arranged based on the value of their index.

def alg_Straight(list_m, list_f):

    PF_pts = Preg_Fund/12
    PE_pts = Preg_Esp/7

    match_score = 0.0

    value_m = 0
    value_f = 0

    #First iterate every client of every list
    for male in list_m:
        for female in list_f:
            #index from 10 to 33 - Preguntas Fundamentales
            for index in range(10, 34, 2):
                value_m = male[index] * male[index+1]
                value_f = female[index] * female[index+1]
                #Checks if theres a match or not on that question/filter and adds points if there is
                if(value_m < 0) and (value_f < 0):
                    match_score += PF_pts
                elif(value_m == 0) and (value_f == 0):
                    match_score += PF_pts
                elif(value_m > 0) and (value_f > 0):
                    match_score += PF_pts

                #Reset values to use them again
                value_m = 0
                value_f = 0

            #index from 34 to 39 - Preguntas Especificas (-Gatos/Perros)
            for index in range(34, 40):
                if (male[index] < 0) and (female[index] < 0):
                    match_score += PE_pts
                elif (male[index] == 0) and (female[index] == 0):
                    match_score += PE_pts
                elif (male[index] > 0) and (female[index] > 0):
                    match_score += PE_pts
            # index for last question - Had bad scale for cats vs dogs
            if(male[40] == 2) and (female[40] == 2):
                match_score += PE_pts
            elif(male[40] == -1) and (female[40] == -1):
                match_score += PE_pts
            elif(male[40] == -2) and (female[40] == -2):
                match_score += PE_pts
            elif (male[40] == 1) and (female[40] == 1):
                match_score += PE_pts
            elif(male[40] == 0) and (female[40] == -2) or (male[40] == 0) and (female[40] == -1):
                match_score += PE_pts
            elif (male[40] == -2) and (female[40] == 0) or (male[40] == -1) and (female[40] == 0):
                match_score += PE_pts


            # Index 41 = added space for match rate
            female[41] = match_score
            match_score = 0.0

        #Organizing an array of lists, from highest to lowest,
        #based on a specific index that is the same for every other list

        index_to_sort = 41

        All_Scores = sorted(list_f, key=lambda x: x[index_to_sort], reverse=True)

        print_Res(male)
        # index 7 is your career
        # index 8 and 9 are the eliminated careers

        count3 = 0
        count2 = 0

        #First printing the top 3 scores, considering their careers to eliminate
        print("Best matches:")
        for w in All_Scores:
            if(male[8] != w[7]) or (male[9] != w[7]):
                print_Partner(male, w)
                count3 += 1
            if(count3 == 3):
                break
        #Second, printing the top 2 scores, getting the excluded careers
        print("Matches that you made, but you excluded ", male[8], " and ", male[9],", their career type:")
        for w in All_Scores:
            if (male[8] == w[7]) or (male[9] == w[7]):
                print_Partner(male, w)
                count2 += 1
            if(count2 == 2):
                break

        #Resets
        for w in All_Scores:
            w[41] = 0
        All_Scores = []

#Using Algorithm for serious-straight couples
print("* * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * ")
print("List of serious-straight couples")
print("--------------------------------------------------------")
print("")
print("//////////////////////////////////////")
print("List for men: ")
alg_Straight(straight_ser_list_m, straight_ser_list_f)
print("--------------------------------------------------------")
print("")
print("//////////////////////////////////////")
print("List for women: ")
alg_Straight(straight_ser_list_f, straight_ser_list_m)
print("--------------------------------------------------------")
print("")

#Algorithm for casual-straight couples
print("* * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * ")
print("List of casual-straight couples")
print("--------------------------------------------------------")
print("")
print("//////////////////////////////////////")
print("List for men: ")
alg_Straight(straight_cas_list_m, straight_cas_list_f)
print("--------------------------------------------------------")
print("")
print("//////////////////////////////////////")
print("List for women: ")
alg_Straight(straight_cas_list_f, straight_cas_list_m)
print("--------------------------------------------------------")
print("")

#Algorithm for gays
print("* * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * ")
print("List for gays:")
alg_Straight(gay_list, gay_list)
print("--------------------------------------------------------")
print("")

#Algorithm for lesb
print("* * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * ")
print("List for lesbians:")
alg_Straight(lesb_list, lesb_list)
print("--------------------------------------------------------")
print("")











































